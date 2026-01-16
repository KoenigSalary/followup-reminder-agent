# -*- coding: utf-8 -*-

import pandas as pd
from datetime import datetime
import os
import warnings
import uuid
from openpyxl import Workbook, load_workbook

warnings.filterwarnings("ignore")


class ExcelHandler:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path

        # Required columns for registry
        self.required_columns = [
            "task_id",
            "meeting_id",
            "Subject",
            "Owner",
            "CC",
            "Due Date",
            "Remarks",
            "Priority",
            "Status",
            "Created On",
            "Last Updated",
            "Last Reminder Date",
            "Last Reminder On",
            "Completed Date",
            "Auto Reply Sent"
        ]

        self._ensure_file_exists()

    # --------------------------------------------------
    # File handling
    # --------------------------------------------------
    def _ensure_file_exists(self):
        """
        Create the Excel file if it doesn't exist, and ensure the header row
        contains the required columns.
        """
        folder = os.path.dirname(self.excel_path)
        if folder:
            os.makedirs(folder, exist_ok=True)

        # Create workbook if missing
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Tasks"
            ws.append(self.required_columns)  # header
            wb.save(self.excel_path)
            return

        # If file exists, ensure required columns exist
        wb = load_workbook(self.excel_path)
        ws = wb["Tasks"] if "Tasks" in wb.sheetnames else wb.active

        # If header row is empty, write it
        header = [cell.value for cell in ws[1] if cell.value is not None]
        if not header:
            ws.append(self.required_columns)
            wb.save(self.excel_path)
            return

        missing = [c for c in self.required_columns if c not in header]
        if missing:
            new_header = header + missing
            for col_idx, col_name in enumerate(new_header, start=1):
                ws.cell(row=1, column=col_idx).value = col_name
            wb.save(self.excel_path)

    # --------------------------------------------------
    # Bulk upload helper
    # --------------------------------------------------
    def add_task(self, task_data: dict) -> None:
        """
        Append one task row to the registry.
        task_data typically contains: Subject, Owner, CC, Due Date, Remarks, Priority, Status
        """
        wb = load_workbook(self.excel_path)
        ws = wb["Tasks"] if "Tasks" in wb.sheetnames else wb.active

        headers = [cell.value for cell in ws[1] if cell.value is not None]
        if not headers:
            headers = self.required_columns
            ws.append(headers)

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        task_data = dict(task_data)
        task_data.setdefault("task_id", str(uuid.uuid4()))
        task_data.setdefault("meeting_id", "")
        task_data.setdefault("Created On", now_str)
        task_data.setdefault("Last Updated", now_str)
        task_data.setdefault("Last Reminder Date", "")
        task_data.setdefault("Last Reminder On", "")
        task_data.setdefault("Completed Date", "")
        task_data.setdefault("Auto Reply Sent", "")

        # Ensure required columns exist in sheet header
        missing = [c for c in self.required_columns if c not in headers]
        if missing:
            headers = headers + missing
            for col_idx, col_name in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx).value = col_name

        ws.append([task_data.get(h, "") for h in headers])
        wb.save(self.excel_path)

    # --------------------------------------------------
    # Read / write as DataFrame
    # --------------------------------------------------
    def load_data(self) -> pd.DataFrame:
        try:
            if not os.path.exists(self.excel_path):
                return pd.DataFrame(columns=self.required_columns)

            df = pd.read_excel(self.excel_path, engine="openpyxl")

            for col in self.required_columns:
                if col not in df.columns:
                    df[col] = None

            return df[self.required_columns]

        except Exception as e:
            print(f"❌ Excel load error: {e}")
            return pd.DataFrame(columns=self.required_columns)

    def save_data(self, df: pd.DataFrame) -> bool:
        try:
            # Ensure order
            for col in self.required_columns:
                if col not in df.columns:
                    df[col] = None
            df = df[self.required_columns]

            df.to_excel(self.excel_path, index=False, engine="openpyxl")
            return True
        except Exception as e:
            print(f"❌ Excel save error: {e}")
            return False

    # --------------------------------------------------
    # Create (manual entry)
    # --------------------------------------------------
    def add_entry(
        self,
        subject: str,
        owner: str,
        due_date,
        remarks: str = "",
        priority: str = "MEDIUM",
        cc: str = "",
        task_id: str = None,
        meeting_id: str = None
    ):
        if not task_id:
            today = datetime.now()
            task_id_prefix = f"MAN-{today.strftime('%Y%m%d')}"
            df = self.load_data()

            if len(df) > 0 and "task_id" in df.columns:
                today_manual_tasks = df[df["task_id"].astype(str).str.startswith(task_id_prefix, na=False)]
                next_seq = len(today_manual_tasks) + 1
            else:
                next_seq = 1

            task_id = f"{task_id_prefix}-{next_seq:03d}"

        if not meeting_id:
            meeting_id = f"MANUAL-{datetime.now().strftime('%Y%m%d')}"

        row = {
            "task_id": task_id,
            "meeting_id": meeting_id,
            "Subject": subject,
            "Owner": owner,
            "CC": cc,
            "Due Date": due_date,
            "Remarks": remarks,
            "Priority": priority,
            "Status": "OPEN",
            "Created On": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Last Updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Last Reminder Date": "",
            "Last Reminder On": "",
            "Completed Date": "",
            "Auto Reply Sent": ""
        }
        return self.append_row(row)

    def append_row(self, row: dict):
        return self.append_rows([row])

    def append_rows(self, rows: list):
        if not rows:
            return 0

        df = self.load_data()
        new_df = pd.DataFrame(rows)

        for col in self.required_columns:
            if col not in new_df.columns:
                new_df[col] = None

        new_df = new_df[self.required_columns]
        combined = pd.concat([df, new_df], ignore_index=True)
        self.save_data(combined)
        return len(combined)

    # --------------------------------------------------
    # Update (index-based)
    # --------------------------------------------------
    def update_row(self, index: int, updates: dict) -> bool:
        try:
            df = self.load_data()
            if index < 0 or index >= len(df):
                return False

            for col, val in updates.items():
                if col in df.columns:
                    df.at[index, col] = val

            df.at[index, "Last Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.save_data(df)
            return True

        except Exception as e:
            print(f"❌ update_row error: {e}")
            return False

    def update_status(self, index: int, status: str) -> bool:
        return self.update_row(index, {"Status": status})

    def delete_row(self, index: int) -> bool:
        return self.update_row(index, {"Status": "DELETED"})

    # --------------------------------------------------
    # Stats
    # --------------------------------------------------
    def get_total_rows(self) -> int:
        return len(self.load_data())
